from openpyxl.styles import Font, NamedStyle, Alignment, Border, Side, PatternFill
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'ADAUSDT'
ws.column_dimensions['A'].width = 17
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.merge_cells('A1:C1')

# ws.row_dimensions[2].height  = 40

rows = [
    ['Account', 'Week', 'Mounth'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]

for row in rows:
    ws.append(row)

header = NamedStyle(name="header")
header.font = Font(bold=True, sz=20)
style_border = Side(border_style="thick", color='0330F5')
header.border = Border(top=style_border, bottom=style_border,
                       right=style_border, left=style_border)
header.alignment = Alignment(horizontal="center", vertical="center")

# Now let's apply this to all first row (header) cells
header_row = ws[1]
for cell in header_row:
    cell.style = header
# style on second header
header_row = ws[2]
for cell in header_row:
    cell.font = Font(bold=True, sz=15)

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      patternType='solid')
greenFill = PatternFill(start_color='00FF00',
                        end_color='00FF00',
                        patternType='solid')

for cell in ws['B'][2:]:
    cell.style = 'Currency'
    if float(cell.value) > 0:
        cell.fill = greenFill
    else:
        cell.fill = redFill
for cell in ws['C'][2:]:
    cell.style = 'Currency'
    if float(cell.value) > 0:
        cell.fill = greenFill
    else:
        cell.fill = redFill

for row in ws:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(filename="t.xlsx")
