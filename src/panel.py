import coloredlogs
import logging
import configparser
from pyrogram.errors import RPCError, BadRequest, NotAcceptable, Unauthorized
from pyrogram.types import ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from pyrogram.enums import ParseMode
from models import *
from binance_api import Binance
from binance.error import ClientError

from main import PRIVATE_LOG_ID as Id_private_log, PUBLIC_LOG_ID as Id_public_log

import random
import string
import datetime

from openpyxl.styles import Font, NamedStyle, Alignment, Border, Side, PatternFill, GradientFill
from openpyxl import Workbook

from apscheduler.schedulers.background import BackgroundScheduler
scheduler = BackgroundScheduler()
scheduler.start()

coloredlogs.install(level='INFO')
logger = logging.getLogger(__name__)

config = configparser.ConfigParser()
config.optionxform = str
config.read('config/bot.ini')

users_data = dict()

B_settings = "تنظیمات ⚙"
B_status_positions_account1 = "وضعیت پوزیشن های اکانت اول 👀"
B_status_positions_accounts = "وضعیت پوزیشن های اکانت ها 👀"
B_status_positions_public = "انتشار وضعیت پوزیشن ها در کانال عمومی ⬅"
B_momentary_balances = "موجودی لحظه ای 💰"
B_pnls = "PNLS"


class MyStartHandler:
    def __init__(self, client, message):
        self.message = message
        self.text = self.message.text
        self.client = client
        self.user_id = self.message.chat.id

        try:
            global users_data
            if self.user_id not in users_data:
                users_data[self.user_id] = {}
            self.user_data = users_data[self.user_id]

            self.run()

        except Exception as e:
            client.send_message(
                self.user_id, repr(e))
            raise e

    def run(self):
        self.user_data.clear()
        text = """به ربات خوش آمدید.🌷"""
        reply_markup = ReplyKeyboardMarkup(
            [
                [B_status_positions_accounts, B_status_positions_account1],
                [B_status_positions_public],
                [B_momentary_balances],
                [B_pnls]

            ],
            resize_keyboard=True)
        self.message.reply(text, reply_markup=reply_markup)


B_volume = "حجم ورود :"


class MyButtonHandler:
    def __init__(self, client, message):
        self.message = message
        self.text = self.message.text
        self.client = client
        self.user_id = self.message.chat.id
        # self.settings = Settings.get(Settings.id == 0)

        try:
            global users_data
            if self.user_id not in users_data:
                users_data[self.user_id] = {}
            self.user_data = users_data[self.user_id]

            self.run()

        except Exception as e:
            client.send_message(
                self.user_id, repr(e))
            raise e

    def run(self):
        if self.text == B_status_positions_account1:
            account_list = config["ACCOUNTS"]["account1"].split(",")
            key = account_list[0]
            secret = account_list[1]
            base_binance = Binance(key=key, secret=secret)

            signals = Signals.select().where(Signals.status == "CLOSE")
            open_position_symbols = [signal.symbol for signal in signals]
            print(open_position_symbols)
            open_position_symbols = list(dict.fromkeys(open_position_symbols))
            if len(open_position_symbols) == 0:
                text = """پوزیشن بازی وجود ندارد!"""
                return self.message.reply(text=text)

            text = ""
            for symbol in open_position_symbols:
                print(symbol)
                # # if signal is not open
                # if len(Targets.select().where((Targets.owner == symbol) and (Targets.status == 'OPEN'))) == 0:
                #     continue

                position = base_binance.get_position(symbol)
                # print(position)
                leverage = int(position['leverage'])
                size = float(position['positionAmt'])
                pnl = round(float(position['unRealizedProfit']), 2)
                liq = round(float(position['liquidationPrice']), 2)
                entry = round(float(position['entryPrice']), 4)
                mark = round(float(position['markPrice']), 2)

                if entry == 0.0:
                    continue

                pnl_percent = round(((mark/entry)-1)*leverage*100, 2)

                margin = (size*mark)/leverage
                margin = round(margin, 4)

                if text != "":
                    text += "➰➰➰➰➰➰➰➰➰➰➰➰"

                text += f"""
📍**{symbol}**      ⚓️**{leverage}**X
❗️PNL:**{pnl}**     ❗️PNL%:**{pnl_percent}**                 
🔸ENTRY:**{entry}**     🔸MARK:**{mark}**
💰MARGIN:**{margin}**       ⚠️Liq:**{liq}**
"""

            if text == "":
                text = """پوزیشن بازی وجود ندارد!"""
            self.message.reply(text=text, parse_mode=ParseMode.MARKDOWN)

        elif self.text == B_status_positions_accounts:
            # signals = Signals.select().where(Signals.status == "CLOSE")
            signals = Signals.select().order_by(Signals.id)
            open_position_symbols = [signal.symbol for signal in signals]
            if len(open_position_symbols) == 0:
                text = """پوزیشن بازی وجود ندارد!"""
                return self.message.reply(text=text)

            open_position_symbols = open_position_symbols[-5:]

            open_position_symbols = [[InlineKeyboardButton(
                symbol, callback_data=f"positions_{symbol}")] for symbol in open_position_symbols]
            reply_markup = InlineKeyboardMarkup(open_position_symbols)

            self.message.reply(text="select one item.",
                               reply_markup=reply_markup)

        elif self.text == B_status_positions_public:
            # signals = Signals.select().where(Signals.status == "CLOSE")
            signals = Signals.select().order_by(Signals.id)
            open_position_symbols = [signal.symbol for signal in signals]
            if len(open_position_symbols) == 0:
                text = """پوزیشن بازی وجود ندارد!"""
                return self.message.reply(text=text)

            open_position_symbols = [open_position_symbols[-1]]

            global text_positions, counter_job
            text_positions = ""
            counter_job = 0
            counter = 0
            for symbol in open_position_symbols:
                text_positions = ""

                accounts = config["ACCOUNTS"].items()
                proxies = list(config["PROXIES"].values())
                div = int(len(accounts)/len(proxies))+1
                proxy = proxies[0]
                logger.info("Connecting with proxy: %s", proxy)
                i = 0

                for account_name, account in accounts:
                    i += 1
                    if (i % div == 0) and i != 1:
                        proxy = proxies[proxies.index(proxy)+1]
                        logger.info("Connecting with proxy: %s", proxy)

                    account_list = account.split(",")
                    key = account_list[0]
                    secret = account_list[1]

                    scheduler.add_job(
                        job_status_positions,
                        args=[self.message, symbol, key,
                              secret, account_name, proxy, True],
                        misfire_grace_time=None
                    )
                    counter += 1

            while counter_job != counter:
                if len(text_positions) >= 3500:
                    self.client.send_message(Id_public_log, text_positions)
                    text_positions = ""
                pass
            # if text_positions == "":
            #     text_positions = """پوزیشن بازی وجود ندارد!"""
            self.client.send_message(Id_public_log, text_positions)

            self.message.reply('☑')

        elif self.text == B_momentary_balances:

            global text_balances, counter_job_balances

            text_balances = ""
            counter_job_balances = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0

            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_get_balances,
                    args=[key, secret, account_name, proxy],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_balances != counter:
                if len(text_balances) >= 3500:
                    self.client.send_message(Id_private_log, text_balances)
                    text_balances = ""

            self.client.send_message(Id_private_log, text_balances)

            self.message.reply('☑')

        elif self.text == B_pnls:
            signals = Signals.select().order_by(Signals.id)
            open_position_symbols = [signal.symbol for signal in signals]
            if len(open_position_symbols) == 0:
                text = """ارزی وجود ندارد!"""
                return self.message.reply(text=text)

            open_position_symbols = open_position_symbols[-5:]

            open_position_symbols = [[InlineKeyboardButton(
                symbol, callback_data=f"pnls_{symbol}")] for symbol in open_position_symbols]
            open_position_symbols.append(
                [InlineKeyboardButton('ALL PNLS', callback_data=f"pnls_ALL")])
            # open_position_symbols = open_position_symbols.append([InlineKeyboardButton('ALL PNLS', callback_data=f"pnls_ALL")])
            reply_markup = InlineKeyboardMarkup(open_position_symbols)

            self.message.reply(text="select one item.",
                               reply_markup=reply_markup)

        elif self.text.startswith('/set_target '):
            if not self.message.reply_to_message:
                text = "لطفا روی سیگنال رپلای کنید!"
                return self.message.reply_text(text)

            target = self.text.replace("/set_target ", "")
            target = float(target)

            # get id_signal from inlineKeyboardButton
            id_signal = self.message.reply_to_message.reply_markup.inline_keyboard[
                0][0].callback_data
            id_signal = id_signal.replace("closetargets_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            # prevent to set terget that may make loss
            account_list = config["ACCOUNTS"]["account1"].split(",")
            key = account_list[0]
            secret = account_list[1]
            base_binance = Binance(key=key, secret=secret)
            price = base_binance.get_price(signal.symbol)
            if signal.kind == 'long':
                if target < price:
                    text = "⚠ Target << Price !"
                    return self.message.reply_text(text)
            else:
                if target > price:
                    text = "⚠ Target >> Price !"
                    return self.message.reply_text(text)

            text = """در حال تنظیم تارگت اصلی ... 🔄"""
            self.message.reply_text(text)

            id_target = randStr()

            global text_settarget, counter_job_settarget
            text_settarget = ""
            counter_job_settarget = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_set_target,
                    args=[key, secret, account_name,
                          proxy, signal, target, id_target],
                    misfire_grace_time=None
                )
                counter += 1

            # Save targets in database
            number_target = 0
            t = Targets.create(
                owner=signal, number=number_target,
                id_target=id_target
            )

            while counter_job_settarget != counter:
                if len(text_settarget) >= 3500:
                    self.client.send_message(Id_private_log, text_settarget)
                    text_settarget = ""

            if text_settarget != "":
                self.client.send_message(Id_private_log, text_settarget)

            text = """تارگت تنظیم شد . ☑️"""
            # print(text)
            self.message.reply(text=text, reply_to_message_id=self.message.id)

        elif self.text.startswith('/limit_balance '):
            limit_balance = self.text.replace("/limit_balance ", "")
            limit_balance = float(limit_balance)

            self.settings.set_limit_balance(limit_balance)

            text = """ تنظیم شد . ☑️"""
            self.message.reply(text=text, reply_to_message_id=self.message.id)

        elif self.text.startswith('/set_stop '):
            if not self.message.reply_to_message:
                text = "لطفا روی سیگنال رپلای کنید!"
                return self.message.reply_text(text)

            stop_loss = self.text.replace("/set_stop ", "")
            stop_loss = float(stop_loss)

            # get id_signal from inlineKeyboardButton
            id_signal = self.message.reply_to_message.reply_markup.inline_keyboard[
                0][0].callback_data
            id_signal = id_signal.replace("closetargets_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            # prevent to set wrong stop loss
            account_list = config["ACCOUNTS"]["account1"].split(",")
            key = account_list[0]
            secret = account_list[1]
            base_binance = Binance(key=key, secret=secret)
            price = base_binance.get_price(signal.symbol)
            if signal.kind == 'long':
                if stop_loss > price:
                    text = "⚠ Stop Loss >> Price !"
                    return self.message.reply_text(text)
            else:
                if stop_loss < price:
                    text = "⚠ Stop Loss << Price !"
                    return self.message.reply_text(text)

            text = """در حال تنظیم استاپ لاس ... 🔄"""
            self.message.reply_text(text)

            id_stop = randStr()+"_stoploss"

            global text_setstop, counter_job_setstop
            text_setstop = ""
            counter_job_setstop = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                     job_set_stop_loss,
                    args=[key, secret, account_name,
                          proxy, signal, stop_loss, id_stop],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_setstop != counter:
                if len(text_setstop) >= 3500:
                    self.client.send_message(Id_private_log, text_setstop)
                    text_setstop = ""

            if text_setstop != "":
                self.client.send_message(Id_private_log, text_setstop)

            text = """استاپ لاس تنظیم شد . ☑️"""
            # print(text)
            self.message.reply(text=text, reply_to_message_id=self.message.id)


def job_status_positions(message, symbol, key, secret, account_name, proxy, public=False):
    global text_positions, counter_job
    # logger.info(f"Getting positions, for account : {account_name} .")
    binance = Binance(key=key, secret=secret, proxy=proxy)

    try:
        position = binance.get_position(symbol)
        # print(position)
        leverage = int(position['leverage'])
        size = float(position['positionAmt'])
        pnl = round(float(position['unRealizedProfit']), 2)
        liq = round(float(position['liquidationPrice']), 2)
        entry = round(float(position['entryPrice']), 4)
        mark = round(float(position['markPrice']), 2)

    except Exception as e:
        print(e)
        counter_job += 1
        return

    if text_positions != "":
        text_positions += "➰➰➰➰➰➰➰➰➰➰➰➰"

    if entry == 0.0:
        text_positions += f"""
📌account : {account_name}  📍**{symbol}**
NOT FOUND ❌
"""

        counter_job += 1
        return

    pnl_percent = round(((mark/entry)-1)*leverage*100, 2)

    margin = (size*mark)/leverage
    margin = round(margin, 4)

    if public:
        text_positions += f"""
📌account : {account_name}
📍**XXXUSDT**      ⚓️**{leverage}**X
❗️PNL:**{pnl}**     ❗️PNL%:**{pnl_percent}**                 
🔸ENTRY:**X.x**     🔸MARK:**X.x**
💰MARGIN:**{margin}**       ⚠️Liq:**{liq}**f
"""
    else:
        text_positions += f"""
📌account : {account_name}
📍**{symbol}**      ⚓️**{leverage}**X
❗️PNL:**{pnl}**     ❗️PNL%:**{pnl_percent}**                 
🔸ENTRY:**{entry}**     🔸MARK:**{mark}**
💰MARGIN:**{margin}**       ⚠️Liq:**{liq}**f
"""

    # print(text_positions)
    counter_job += 1


def job_get_balances(key, secret, account_name, proxy):
    global text_balances, counter_job_balances

    if text_balances == "":
        text_balances += "**💰All Balances💰**\n\n"

    try:
        binance = Binance(key=key, secret=secret, proxy=proxy)
        balance = binance.get_balance()
        # logger.info(f"Balance for account : {key[:20]}, is : {balance}")
        text_balances += f"📌account : {account_name}\n💲balance : **{balance}**\n\n"
    except Exception as error:
        print(error)

    counter_job_balances += 1


def job_get_pnls(symbol, start_time_lastweek_timestamp, start_time_lastmounth_timestamp,
                 end_time_timestamp, key, secret, account_name, proxy):
    global text_pnls, counter_job_pnls

    # if text_pnls == "":
    #     text_pnls += "**💰All PNLS**\n\n"

    try:
        binance = Binance(key=key, secret=secret, proxy=proxy)
        pnl_lastday = binance.get_last_pnl(
            symbol, end_time_timestamp, end_time_timestamp)
        pnl_lastday = round(pnl_lastday, 2)
        pnl_lastweek = binance.get_last_pnl(
            symbol, start_time_lastweek_timestamp, end_time_timestamp)
        pnl_lastweek = round(pnl_lastweek, 2)
        pnl_lastmounth = 0
        try:
            pnl_lastmounth = binance.get_last_pnl(
                symbol, start_time_lastmounth_timestamp, end_time_timestamp)
            pnl_lastmounth = round(pnl_lastmounth, 2)
        except Exception as error:
            print(error)
        # logger.info(f"Balance for account : {key[:20]}, is : {balance}")
        # text_pnls += f"📌account : {account_name}\npnl : **{pnl}**\n\n"
        text_pnls += f"\n{account_name},{pnl_lastday},{pnl_lastweek},{pnl_lastmounth}"

    except Exception as error:
        print(error)

    counter_job_pnls += 1


def job_set_target(key, secret, account_name, proxy, signal, target, id_target):
    global text_settarget, counter_job_settarget
    # logger.info(f"Cancelling signal from user, for account : {account_name} .")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    openOrder = binance.get_order(
        symbol=signal.symbol, ClientOrderId=signal.id_signal)
    if not openOrder:
        text = f'order not setted for account {account_name} yet!'
        logger.warn(text)
        text_settarget += text + "\n"
        counter_job_settarget += 1
        return
    size = float(openOrder['origQty'])

    decimal_coin = binance.get_decimal_coin(signal.symbol)
    price_target = target
    percent_target = 100

    size_target = (size*percent_target)/100
    size_target = float(format(size_target, f'.{decimal_coin}f'))

    logging.info(
        f"Setting target: {price_target} with size: {size_target}, for symbol: {signal.symbol}")
    try:
        if signal.kind == 'long':
            order = binance.limit_short(
                signal.symbol, price_target, size_target, ClientOrderId=id_target)
        else:
            order = binance.limit_long(
                signal.symbol, price_target, size_target, ClientOrderId=id_target)
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
        text = f"""
**🚨Log in setting targets.**
**Account** : {account_name}
**Error :** `{text}`            """
        text_settarget += text+'\n'
        # bot.send_message(admin, text)

    counter_job_settarget += 1


def job_set_stop_loss(key, secret, account_name, proxy, signal, stop_loss, id_stop):
    global text_setstop, counter_job_setstop
    # logging.info(f"Setting stop loss with hand ...")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    # stop_loss
    try:
        ClientOrderId = id_stop
        print(ClientOrderId)
        if signal.kind == 'long':
            order = binance.stoploss_short(
                signal.symbol, stop_loss, ClientOrderId)
        else:
            order = binance.stoploss_long(
                signal.symbol, stop_loss, ClientOrderId)
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)

        text = f"""
**🚨Log in setting stop loss.**
**Account** : {account_name}
**Error :** `{text}`            """

        text_setstop += text+'\n'
        counter_job_setstop += 1

        return

    # save clientOrderId and orderId stoploss for cancel it later
    signal.set_id_stoploss(order['orderId'])
    signal.set_client_id_stoploss(order['clientOrderId'])
    counter_job_setstop += 1


B_back = "برگشت 🔙"


class MyCallbackHandler:
    def __init__(self, client, callback_query):
        self.client = client
        self.query = callback_query
        self.message = self.query.message
        self.user_id = self.query.from_user.id
        self.name = self.query.from_user.first_name
        self.data = self.query.data
        # self.settings = Settings.get(Settings.id == 0)

        try:
            logger.info("get callBack.")
            self.run()
            client.answer_callback_query(self.query.id)
        except BadRequest:
            text = "خطا!"
            client.answer_callback_query(self.query.id, text)
        except Exception as e:
            client.send_message(
                self.user_id, repr(e))
            raise e

    def run(self):

        if self.data.startswith("cancel_"):
            text = """در حال انجام ... 🔄"""
            self.client.answer_callback_query(
                self.query.id, text=text)

            id_signal = self.data.replace("cancel_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            global text_closepositions, counter_job_closepositions
            text_closepositions = ""
            counter_job_closepositions = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_close_position,
                    args=[key, secret, account_name, proxy, signal],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_closepositions != counter:
                if len(text_closepositions) >= 3500:
                    self.client.send_message(
                        Id_private_log, text=text_closepositions, reply_to_message_id=self.message.id)
                    text_closepositions = ""
            # print(counter_job_closepositions,counter)

            if text_closepositions != "":
                self.client.send_message(
                    Id_private_log, text=text_closepositions,
                    parse_mode=ParseMode.MARKDOWN, reply_to_message_id=self.message.id)

            signal.set_status('CANCELED')

            text = """پوزیشن بسته شد . ☑️"""
            # print(text)
            self.message.reply(text=text, reply_to_message_id=self.message.id)
            # self.client.answer_callback_query(
            #     self.query.id, text=text, show_alert=True)

            # self.client.edit_message_reply_markup(
            #     chat_id=self.message.chat.id,
            #     message_id=self.message.id,
            #     reply_markup=None)

        elif self.data.startswith("closestop_"):
            text = """در حال انجام ... 🔄"""
            self.client.answer_callback_query(
                self.query.id, text=text)

            id_signal = self.data.replace("closestop_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            global text_closestop, counter_job_closestop
            text_closestop = ""
            counter_job_closestop = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_close_stop_loss,
                    args=[key, secret, account_name, proxy, signal],
                    misfire_grace_time=None
                )
                counter += 1

                # pass close stop loss on other action
                # signal.set_client_id_stoploss('a')

            while counter_job_closestop != counter:
                if len(text_closestop) >= 3500:
                    self.client.send_message(
                        Id_private_log, text=text_closestop,
                        reply_to_message_id=self.message.id)
                    text_closestop = ""

            if text_closestop != "":
                self.client.send_message(
                    Id_private_log, text=text_closestop,
                    parse_mode=ParseMode.MARKDOWN, reply_to_message_id=self.message.id)

            text = """استاپ لاس بسته شد . ☑️"""
            self.client.send_message(
                self.user_id, text=text, reply_to_message_id=self.message.id)
            # self.client.answer_callback_query(
            #     self.query.id, text=text, show_alert=True)

            reply_markup = InlineKeyboardMarkup(
                [[InlineKeyboardButton("❌ cancel ❌", callback_data=f"cancel_{id_signal}")]])

            # self.client.edit_message_reply_markup(
            #     chat_id=self.message.chat.id,
            #     message_id=self.message.id,
            #     reply_markup=reply_markup)

        elif self.data.startswith("closetargets_"):
            text = """در حال انجام ... 🔄"""
            self.client.answer_callback_query(
                self.query.id, text=text)

            id_signal = self.data.replace("closetargets_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            global text_closetargets, counter_job_closetargets
            text_closetargets = ""
            counter_job_closetargets = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_close_targets,
                    args=[key, secret, account_name, proxy, signal],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_closetargets != counter:
                if len(text_closetargets) >= 3500:
                    self.client.send_message(
                        Id_private_log, text=text_closetargets,
                        reply_to_message_id=self.message.id)
                    text_closetargets = ""

            if text_closetargets != "":
                self.client.send_message(
                    Id_private_log, text=text_closetargets,
                    parse_mode=ParseMode.MARKDOWN, reply_to_message_id=self.message.id)

            text = """تارگت ها بسته شد . ☑️"""
            self.client.send_message(
                self.user_id, text=text, reply_to_message_id=self.message.id)
            # self.client.answer_callback_query(
            #     self.query.id, text=text, show_alert=True)

        elif self.data.startswith("rollingstop_"):
            text = """در حال انجام ... 🔄"""
            self.client.answer_callback_query(
                self.query.id, text=text)

            id_signal = self.data.replace("rollingstop_", "")
            signal = Signals.get(Signals.id_signal == id_signal)

            account_list = config["ACCOUNTS"]["account1"].split(",")
            key = account_list[0]
            secret = account_list[1]
            base_binance = Binance(key=key, secret=secret)
            price = base_binance.get_price(signal.symbol)

            entry = signal.entry
            if 'market' in str(entry):
                entry = float(signal.entry.replace('market', ''))
            if signal.kind == 'long':
                if entry > price:
                    text = "⚠ Entry >> Price !"
                    return self.message.reply_text(text)
            else:
                if entry < price:
                    text = "⚠ Entry << Price !"
                    return self.message.reply_text(text)

            global text_rollingstop, counter_job_rollingstop
            text_rollingstop = ""
            counter_job_rollingstop = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_rolling_stop_loss,
                    args=[key, secret, account_name, proxy, signal],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_rollingstop != counter:
                if len(text_rollingstop) >= 3500:
                    self.client.send_message(
                        Id_private_log, text=text_rollingstop,
                        reply_to_message_id=self.message.id)
                    text_rollingstop = ""

            if text_rollingstop != "":
                self.client.send_message(
                    Id_private_log, text=text_rollingstop,
                    parse_mode=ParseMode.MARKDOWN, reply_to_message_id=self.message.id)

            text = """استاپ لاس روی نقطه ورود تنظیم شد . ☑️"""
            self.client.send_message(
                self.user_id, text=text, reply_to_message_id=self.message.id)

            reply_markup = InlineKeyboardMarkup(
                [[InlineKeyboardButton("❌ cancel ❌", callback_data=f"cancel_{id_signal}")]])

            # self.client.edit_message_reply_markup(
            #     chat_id=self.message.chat.id,
            #     message_id=self.message.id,
            #     reply_markup=reply_markup)

        elif self.data.startswith('positions_'):
            symbol = self.data.replace('positions_', '')
            self.message.delete()

            global text_positions, counter_job
            text_positions = ""
            counter_job = 0
            counter = 0

            print(symbol)
            # # if signal is not open
            # if len(Targets.select().where((Targets.owner == symbol) and (Targets.status == 'OPEN'))) == 0:
            #     continue

            text_positions = ""

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0

            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_status_positions,
                    args=[self.message, symbol, key,
                          secret, account_name, proxy],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job != counter:
                if len(text_positions) >= 3500:
                    self.client.send_message(Id_private_log, text_positions)
                    text_positions = ""
                pass
            # if text_positions == "":
            #     text_positions = """پوزیشن بازی وجود ندارد!"""
            self.client.send_message(Id_private_log, text_positions)

            self.message.reply('☑')

        # elif self.data == B_back:
        #     reply_markup = reply_markup_settings()

        #     self.client.edit_message_reply_markup(
        #         chat_id=self.message.chat.id,
        #         message_id=self.message.message_id,
        #         reply_markup=reply_markup)

        elif self.data.startswith('pnls_'):
            symbol = self.data.replace('pnls_', '')
            if symbol == 'ALL':
                symbol = None
            self.message.delete()

            end_time = datetime.datetime.utcnow()
            end_time_timestamp = int(end_time.timestamp())*1000

            start_time_lastweek = end_time+datetime.timedelta(-7)
            start_time_lastweek_timestamp = int(
                start_time_lastweek.timestamp())*1000
            start_time_lastmounth = end_time+datetime.timedelta(-30)
            start_time_lastmounth_timestamp = int(
                start_time_lastmounth.timestamp())*1000

            global text_pnls, counter_job_pnls

            text_pnls = ""
            counter_job_pnls = 0
            counter = 0

            accounts = config["ACCOUNTS"].items()
            proxies = list(config["PROXIES"].values())
            div = int(len(accounts)/len(proxies))+1
            proxy = proxies[0]
            logger.info("Connecting with proxy: %s", proxy)
            i = 0
            for account_name, account in accounts:
                i += 1
                if (i % div == 0) and i != 1:
                    proxy = proxies[proxies.index(proxy)+1]
                    logger.info("Connecting with proxy: %s", proxy)

                account_list = account.split(",")
                key = account_list[0]
                secret = account_list[1]

                scheduler.add_job(
                    job_get_pnls,
                    args=[symbol, start_time_lastweek_timestamp, start_time_lastmounth_timestamp,
                          end_time_timestamp, key, secret, account_name, proxy],
                    misfire_grace_time=None
                )
                counter += 1

            while counter_job_pnls != counter:
                pass

            rows_exel = [['Account', 'Day', 'Week', 'Month']]
            for account_pnl in text_pnls.split("\n"):
                if not account_pnl:
                    continue
                rows_exel.append(account_pnl.split(','))

            wb = Workbook()
            ws = wb.active

            if not symbol:
                symbol = 'ALL'
            ws['A1'] = f"PNLS {symbol}"
            ws.column_dimensions['A'].width = 17
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.merge_cells('A1:D1')

            for row in rows_exel:
                ws.append(row)

            header = NamedStyle(name="header")
            header.font = Font(bold=True, sz=20)
            style_border = Side(border_style="double", color='b3b005')
            header.border = Border(top=style_border, bottom=style_border,
                                   right=style_border, left=style_border)
            header.alignment = Alignment(
                horizontal="center", vertical="center")

            # Now let's apply this to all first row (header) cells
            header_row = ws[1]
            for cell in header_row:
                cell.style = header
            # style on second header
            header_row = ws[2]
            for cell in header_row:
                cell.font = Font(bold=True, sz=15)

            # redFill = PatternFill(start_color='FFFF0000',
            #                       end_color='FFFF0000',
            #                       patternType='solid')
            redFill = GradientFill(stop=("db1304", "FFFF0000"))
            # greenFill = PatternFill(start_color='00FF00',
            #                         end_color='00FF00',
            #                         patternType='solid')
            greenFill = GradientFill(stop=("02db02", "00FF00"))

            for cell in ws['B'][2:]:
                cell.style = 'Currency'
                if not cell.value:
                    continue
                if float(cell.value) > 0:
                    cell.fill = greenFill
                elif float(cell.value) == 0:
                    pass
                else:
                    cell.fill = redFill
            for cell in ws['C'][2:]:
                cell.style = 'Currency'
                if not cell.value:
                    continue
                if float(cell.value) > 0:
                    cell.fill = greenFill
                elif float(cell.value) == 0:
                    pass
                else:
                    cell.fill = redFill
            for cell in ws['D'][2:]:
                cell.style = 'Currency'
                if not cell.value:
                    continue
                if float(cell.value) > 0:
                    cell.fill = greenFill
                elif float(cell.value) == 0:
                    pass
                else:
                    cell.fill = redFill

            for row in ws:
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")

            wb.save(filename="pnls.xlsx")
            self.client.send_document(
                chat_id=self.user_id, document="pnls.xlsx",)


def job_close_position(key, secret, account_name, proxy, signal):
    global text_closepositions, counter_job_closepositions
    logger.info(f"Cancelling signal from user, for account : {account_name} .")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    # try to cancel order if not open yet
    try:
        openOrder = binance.get_order(
            symbol=signal.symbol, ClientOrderId=signal.id_signal)
        logger.info(openOrder)
        if openOrder:
            binance.cancel_open_order(
                signal.symbol,
                ClientOrderId=openOrder['clientOrderId'],
                orderId=openOrder['orderId']
            )
    except ClientError as error:
        pass

    try:
        binance.cancel_order(signal.symbol, signal.kind)
    except Exception as e:
        pass

    # Get old order stop_loss
    try:
        old_order_stoploss = binance.get_order(
            symbol=signal.symbol, ClientOrderId=signal.client_id_stoploss)
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
        text = f"""\n
**🚨Log in canceling with hand.**
**Account** : {account_name}
**Error :** `{text}`            """
        text_closepositions += text
        # self.client.send_message(self.user_id, text)
        counter_job_closepositions += 1
        return
    # if not found order or None order
    if not old_order_stoploss:
        counter_job_closepositions += 1
        return
    # cancel stoploss
    try:
        binance.cancel_open_order(
            signal.symbol,
            ClientOrderId=old_order_stoploss['clientOrderId'],
            orderId=old_order_stoploss['orderId']
        )
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
        text = f"""\n
**🚨Log in canceling stoploss with hand.**
**Account** : {account_name}
**Error :** `{text}`            """
        text_closepositions += text
        # self.client.send_message(self.user_id, text)

    counter_job_closepositions += 1
    # print('counter_job_closepositions', counter_job_closepositions)


def job_close_stop_loss(key, secret, account_name, proxy, signal):
    global text_closestop, counter_job_closestop
    logger.info(
        f"Clossing stop loss from user, for account : {account_name} .")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    # Get old order stop_loss
    try:
        old_order_stoploss = binance.get_order(
            symbol=signal.symbol, ClientOrderId=signal.client_id_stoploss)
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
#         text = f"""
# **🚨Log in canceling stoploss with hand.**
# **Account** : {account_name}
# **Error :** `{text}`            """
#         self.client.send_message(self.user_id, text)
        # text_closestop += text
        counter_job_closestop += 1
        return
    # if not found order or None order
    if not old_order_stoploss:
        counter_job_closestop += 1
        return
    logger.info(old_order_stoploss)
    # cancel stoploss
    try:
        binance.cancel_open_order(
            signal.symbol,
            ClientOrderId=old_order_stoploss['clientOrderId'],
            orderId=old_order_stoploss['orderId']
        )
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
        text = f"""\n
**🚨Log in canceling stoploss with hand.**
**Account** : {account_name}
**Error :** `{text}`            """
        text_closestop += text
        # self.client.send_message(self.user_id, text)

    counter_job_closestop += 1


def job_close_targets(key, secret, account_name, proxy, signal):
    global text_closetargets, counter_job_closetargets
    logger.info(
        f"Closing targets from user, for account : {account_name} .")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    targets = Targets.select().where(Targets.owner == signal)
    for target in targets:

        try:
            openOrder = binance.get_order(
                symbol=target.owner.symbol, ClientOrderId=target.id_target)
            # logger.info(openOrder)
        except ClientError as error:
            logger.error(
                "Found error. status: {}, error code: {}, error message: {}".format(
                    error.status_code, error.error_code, error.error_message
                )
            )
            continue
        # if not found order or None order
        if not openOrder:
            logger.info(
                f"Target None deleted, for account : {account_name}.")
            continue
        # logger.info(openOrder['status'])

        if openOrder['status'] == 'FILLED':
            continue

        try:
            binance.cancel_open_order(
                target.owner.symbol,
                ClientOrderId=openOrder['clientOrderId'],
                orderId=openOrder['orderId']
            )
            logger.info(
                f"target {openOrder['price']} cancelled, for account : {account_name} .")
        except ClientError as error:
            text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
                error.status_code, error.error_code, error.error_message, account_name
            )
            logger.error(text)

    counter_job_closetargets += 1


def job_rolling_stop_loss(key, secret, account_name, proxy, signal):
    global text_rollingstop, counter_job_rollingstop
    logger.info(
        f"Rolling stop loss from user, for account : {account_name} .")

    binance = Binance(key=key, secret=secret, proxy=proxy)

    # Get old order stop_loss
    try:
        old_order_stoploss = binance.get_order(
            symbol=signal.symbol, ClientOrderId=signal.client_id_stoploss)
    except ClientError as error:
        text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
            error.status_code, error.error_code, error.error_message, account_name
        )
        logger.error(text)
        # counter_job_rollingstop += 1
        # return
    # if not found order or None order
    if old_order_stoploss:
        # counter_job_rollingstop += 1
        # return
        logger.info(old_order_stoploss)
        # cancel stoploss
        try:
            binance.cancel_open_order(
                signal.symbol,
                ClientOrderId=old_order_stoploss['clientOrderId'],
                orderId=old_order_stoploss['orderId']
            )
        except ClientError as error:
            text = "Found error. status: {}, error code: {}, error message: {}, for account: {}".format(
                error.status_code, error.error_code, error.error_message, account_name
            )
            logger.error(text)
            text = f"""\n
    **🚨Log in canceling stoploss with hand.**
    **Account** : {account_name}
    **Error :** `{text}`            """
            text_rollingstop += text
            # self.client.send_message(self.user_id, text)

    openOrder = binance.get_order(
        symbol=signal.symbol, ClientOrderId=signal.id_signal)
    if not openOrder:
        text = f'order not setted for account {account_name} yet!'
        logger.warn(text)
        text_rollingstop += text + "\n"
        counter_job_rollingstop += 1
        return
    print(openOrder)

    stop_loss = float(openOrder['avgPrice'])
    if stop_loss == 0.0:
        stop_loss = float(openOrder['price'])
    print("stop", stop_loss)

    # set stop loss on entry point
    ClientOrderId = signal.id_signal+"_stoploss"+str(int(stop_loss))
    # stop_loss
    if signal.kind == 'long':
        order_stoploss = binance.stoploss_short(
            signal.symbol, stop_loss, ClientOrderId)
    else:
        order_stoploss = binance.stoploss_long(
            signal.symbol, stop_loss, ClientOrderId)

    # save clientOrderId and orderId stoploss for cancel it later
    signal.set_id_stoploss(order_stoploss['orderId'])
    signal.set_client_id_stoploss(
        order_stoploss['clientOrderId'])

    counter_job_rollingstop += 1


def randStr(chars=string.ascii_uppercase + string.ascii_lowercase + string.digits, N=22):
    return ''.join(random.choice(chars) for _ in range(N))
